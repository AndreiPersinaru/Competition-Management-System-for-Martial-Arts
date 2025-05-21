from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.http import quote
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

from ..serializers import CompetitieSerializer
from ..models import Competitie, Sportiv, Club, Proba, Categorie, Inscriere, ClasamentProba, ClasamentClub
from ..utils.categorii import populate_categorii_standard
from ..utils.bracket_generation import genereaza_bracket_si_meciuri
from ..utils.clasament_club import actualizeaza_clasament_cluburi_pentru_competitie

class CompetitieViewSet(viewsets.ModelViewSet):
    queryset = Competitie.objects.all()
    serializer_class = CompetitieSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_authenticators(self):
        if self.request.method == 'GET':
            return []
        return [JWTAuthentication()]

    def perform_create(self, serializer):
        competitie = serializer.save()
        probe_nume = self.request.data.get("probe", [])
        for nume in probe_nume:
            proba, _ = Proba.objects.get_or_create(nume=nume)
            competitie.probe.add(proba)


# ========== DOWNLOAD EXCEL TEMPLATE ==========

class CompetitionExcelTemplateView(APIView):
    permission_classes = []

    def get(self, request, competition_id):
        wb = Workbook()
        competitie = get_object_or_404(Competitie, id=competition_id)
        probe = list(competitie.probe.values_list('nume', flat=True))

        ws = wb.active
        ws.title = "Sportivi"

        headers = ["NR LEG", "NUME SI PRENUME", "CLUB", "GEN", "CNP", "DATA NASTERII", "CATEGORIE VARSTA", "KG", "CATEGORIE KG", "PROBA", "MECI", "TAXA"]
        ws.append(headers)

        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = len(header) + 7
            for row in range(1, 1001):
                ws[f"{col_letter}{row}"].font = openpyxl.styles.Font(bold=True, name="Calibri1")
                ws.row_dimensions[row].height = 25
                ws[f"{col_letter}{row}"].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                

        # Validare Gen
        dv_gen = DataValidation(type="list", formula1='"Masculin,Feminin"', allow_blank=False)
        ws.add_data_validation(dv_gen)
        dv_gen.add("D2:D1000")

        # Validare Proba
        if probe:
            print(f"Probe: {probe}")
            probe_list_str = ','.join([f'{p}' for p in probe])
            dv_probe = DataValidation(type="list", formula1='"' + probe_list_str + '"', allow_blank=False)
            ws.add_data_validation(dv_probe)
            dv_probe.add("J2:J1000")
    
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"Inscriere {competitie.nume}.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        wb.save(response)
        return response
    

# ========== EXPORT PARTICIPANTS ==========
class ExportParticipantsView(APIView):
    permission_classes = []

    def get(self, request, competition_id):
        competitie = get_object_or_404(Competitie, id=competition_id)
        inscrieri = Inscriere.objects.filter(competitie=competitie).select_related("sportiv", "categorie__proba", "sportiv__club")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sportivi"

        headers = ["NR LEG", "NUME SI PRENUME", "CLUB", "GEN", "CNP", "DATA NASTERII", "CATEGORIE VARSTA", "CATEGORIE KG", "PROBA", "MECI", "TAXA"]
        ws.append(headers)

        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = len(header) + 7
            for row in range(1, 1001):
                ws[f"{col_letter}{row}"].font = openpyxl.styles.Font(bold=True, name="Calibri1")
                ws.row_dimensions[row].height = 25
                ws[f"{col_letter}{row}"].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        row_num = 2
        for inscriere in inscrieri:
            sportiv = inscriere.sportiv
            categorie = inscriere.categorie
            if any(p.lower() in categorie.proba.nume.lower() for p in ["Polydamas", "Palaismata"]):
                taxa = 75
            else:
                taxa = 100
            values = [
                sportiv.nr_legitimatie,
                f"{sportiv.nume} {sportiv.prenume}",
                sportiv.club.nume if sportiv.club else "",
                sportiv.sex,
                sportiv.cnp,
                sportiv.data_nastere.strftime("%d.%m.%Y"),
                f"{categorie.varsta_min}-{categorie.varsta_max}",
                f"{categorie.categorie_greutate} KG" if categorie.categorie_greutate else "",
                categorie.proba.nume,
                inscriere.meci_demonstrativ,
                taxa,
            ]
            for col_num, value in enumerate(values, start=1):
                ws.cell(row=row_num, column=col_num).value = value
            row_num += 1

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"Lista Sportivi {competitie.nume}.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        wb.save(response)
        return response
    

# ========== UPLOAD PARTICIPANTS ==========
    
class UploadParticipantsView(APIView):
    permission_classes = []

    def post(self, request, competition_id):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Fisierul nu a fost trimis."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"Fisier Excel invalid: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        competitie = get_object_or_404(Competitie, id=competition_id)
        errors = []
        added_count = 0

        header = [cell.value for cell in ws[1]]
        try:
            col_nr_leg = header.index("NR LEG")
            col_nume_si_prenume = header.index("NUME SI PRENUME")
            col_club = header.index("CLUB")
            col_gen = header.index("GEN")
            col_cnp = header.index("CNP")
            col_data_nasterii = header.index("DATA NASTERII")
            col_cat_kg = header.index("CATEGORIE KG")
            col_proba = header.index("PROBA")
            col_meci_demonstrativ = header.index("MECI")
        except ValueError as e:
            return Response({"detail": f"Header lipsa: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue

                try:
                    nume_prenume = str(row[col_nume_si_prenume]).strip()
                    nume, prenume = nume_prenume.split(" ", 1)
                except Exception:
                    continue

                nr_leg = row[col_nr_leg]
                club_nume = row[col_club]
                sex = row[col_gen]
                if sex == "Masculin":
                    sex = "M"
                elif sex == "Feminin":
                    sex = "F"
                cnp = row[col_cnp]
                data_nastere = row[col_data_nasterii]
                cat_greutate = row[col_cat_kg]
                proba_nume = row[col_proba]
                meci_demonstrativ = row[col_meci_demonstrativ]

                if not all([nr_leg, nume, prenume, club_nume, sex, cnp, data_nastere, proba_nume]):
                    continue

                try:
                    data_nastere = datetime.strptime(str(data_nastere), "%d.%m.%Y").date()
                except Exception:
                    continue

                if len(str(cnp)) != 13:
                    continue

                club, _ = Club.objects.get_or_create(nume=club_nume)

                sportiv, _ = Sportiv.objects.get_or_create(cnp=cnp, defaults={
                    "nume": nume,
                    "prenume": prenume,
                    "sex": sex,
                    "club": club,
                    "data_nastere": data_nastere,
                    "nr_legitimatie": nr_leg,
                })
                
                varsta = (competitie.data_incepere - data_nastere).days // 365


                proba_obj, _ = Proba.objects.get_or_create(nume=proba_nume)

                if not Categorie.objects.filter(proba=proba_obj).exists():
                    populate_categorii_standard(proba_obj)

                # Tratament special pentru Polydamas / Palaismata
                if any(p.lower() in proba_obj.nume.lower() for p in ["Polydamas", "Palaismata"]):
                    categorie = Categorie.objects.filter(
                        proba=proba_obj,
                        varsta_min__lte=varsta,
                        varsta_max__gte=varsta
                    ).first()
                else:
                    def parse_categorie_greutate(raw_kg):
                        try:
                            raw_kg = str(raw_kg).strip().replace(" ", "").replace("KG", "").replace("kg", "")
                            return raw_kg
                        except Exception:
                            return None

                    cat_greutate = parse_categorie_greutate(cat_greutate)
                    if cat_greutate is None:
                        continue

                    categorie = Categorie.objects.filter(
                        proba=proba_obj,
                        sex=sex,
                        varsta_min__lte=varsta,
                        varsta_max__gte=varsta,
                        categorie_greutate=cat_greutate
                    ).first()

                if not categorie:
                    continue

                if not Inscriere.objects.filter(
                    sportiv=sportiv,
                    categorie=categorie,
                    competitie=competitie
                ).exists():
                    Inscriere.objects.create(
                        sportiv=sportiv,
                        categorie=categorie,
                        competitie=competitie,
                        varsta=varsta,
                        meci_demonstrativ=meci_demonstrativ,
                    )
                added_count += 1

            except Exception as e:
                print(f"Randul {idx}: {str(e)}")

        # Get a list of all inscrieri for the competition
        inscrieri = Inscriere.objects.filter(competitie=competitie).select_related("sportiv", "categorie__proba", "sportiv__club")
        genereaza_bracket_si_meciuri(inscrieri, competitie) 

        return Response({
            "detail": f"{added_count} inscrieri adaugate",
        }, status=status.HTTP_201_CREATED)

# ========== DOWNLOAD RANKING EXCEL ==========


class DownloadRankingView(APIView):
    permission_classes = []

    def get(self, request, competition_id):
        competitie = get_object_or_404(Competitie, id=competition_id)
        
        actualizeaza_clasament_cluburi_pentru_competitie(competitie)
        
        # Obținem toate clasamentele pentru această competiție
        clasamente = ClasamentProba.objects.filter(
            competitie=competitie
        ).select_related(
            'sportiv', 'sportiv__club', 'categorie', 'categorie__proba'
        ).order_by('categorie__proba__nume', 'categorie__sex', 'categorie__categorie_greutate', 'puncte')

        # Obținem clasamentul cluburilor
        clasament_cluburi = ClasamentClub.objects.filter(
            competitie=competitie
        ).select_related('club').order_by('-puncte')

        wb = Workbook()
        
        # ============= SHEET 1: CLASAMENT SPORTIVI =============
        ws_sportivi = wb.active
        ws_sportivi.title = "Clasament Sportivi"

        # Headers pentru Excel
        headers_sportivi = ["LOCUL", "NUME SI PRENUME", "PROBA", "GEN", "CATEGORIE KG", "CATEGORIE VARSTA", "CLUB"]
        ws_sportivi.append(headers_sportivi)

        # Styling pentru headers și rânduri
        for i, header in enumerate(headers_sportivi, start=1):
            col_letter = get_column_letter(i)
            ws_sportivi.column_dimensions[col_letter].width = len(header) + 10
            
            # Styling pentru toate rândurile (inclusiv header)
            for row in range(1, 1001):
                cell = ws_sportivi[f"{col_letter}{row}"]
                cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
                ws_sportivi.row_dimensions[row].height = 25
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # Grupăm clasamentele pe categorii pentru a genera locuri corecte
        categorii_clasamente = {}
        for clasament in clasamente:
            categorie_key = (
                clasament.categorie.proba.nume,
                clasament.categorie.sex,
                clasament.categorie.categorie_greutate or '',
                f"{clasament.categorie.varsta_min}-{clasament.categorie.varsta_max}"
            )
            
            if categorie_key not in categorii_clasamente:
                categorii_clasamente[categorie_key] = []
            categorii_clasamente[categorie_key].append(clasament)

        row_num = 2
        
        # Pentru fiecare categorie, sortăm și atribuim locurile
        for categorie_key, clasamente_categorie in categorii_clasamente.items():
            # Sortăm după puncte (locul în clasament)
            clasamente_categorie.sort(key=lambda x: x.puncte)
            
            # Adăugăm rând gol între categorii (cu excepția primei)
            if row_num > 2:
                row_num += 1
            
            # Pentru fiecare sportiv din categorie
            for loc, clasament in enumerate(clasamente_categorie, 1):
                sportiv = clasament.sportiv
                categorie = clasament.categorie
                
                # Determinăm genul pentru afișare
                gen_display = "Masculin" if categorie.sex == "M" else "Feminin"
                
                # Determinăm categoria de greutate
                cat_kg_display = f"{categorie.categorie_greutate} KG" if categorie.categorie_greutate else "N/A"
                
                values = [
                    loc,  # Locul în categorie
                    f"{sportiv.nume} {sportiv.prenume}",
                    categorie.proba.nume,
                    gen_display,
                    cat_kg_display,
                    f"{categorie.varsta_min}-{categorie.varsta_max}",
                    sportiv.club.nume if sportiv.club else ""
                ]
                
                # Adăugăm valorile în Excel
                for col_num, value in enumerate(values, start=1):
                    cell = ws_sportivi.cell(row=row_num, column=col_num)
                    cell.value = value
                    
                    # Font diferit pentru rândurile cu date (nu bold)
                    cell.font = openpyxl.styles.Font(bold=False, name="Calibri")
                
                row_num += 1

        # Dacă nu avem clasamente, adăugăm un mesaj
        if not clasamente:
            ws_sportivi.cell(row=2, column=1).value = "Nu există clasamente disponibile pentru această competiție"
            ws_sportivi.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True, name="Calibri")

        # ============= SHEET 2: CLASAMENT CLUBURI =============
        ws_cluburi = wb.create_sheet(title="Clasament Cluburi")

        # Headers pentru clasamentul cluburilor
        headers_cluburi = ["LOCUL", "CLUB", "PUNCTE"]
        ws_cluburi.append(headers_cluburi)

        # Styling pentru headers și rânduri cluburi
        for i, header in enumerate(headers_cluburi, start=1):
            col_letter = get_column_letter(i)
            ws_cluburi.column_dimensions[col_letter].width = len(header) + 15
            
            # Styling pentru toate rândurile (inclusiv header)
            for row in range(1, 101):  # Reducem numărul de rânduri pre-stilizate pentru sheet-ul de cluburi
                cell = ws_cluburi[f"{col_letter}{row}"]
                cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
                ws_cluburi.row_dimensions[row].height = 25
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # Adăugăm datele pentru clasamentul cluburilor
        row_num_cluburi = 2
        for loc, clasament_club in enumerate(clasament_cluburi, 1):
            values_cluburi = [
                loc,  # Locul în clasament
                clasament_club.club.nume,
                clasament_club.puncte,
            ]
            
            # Adăugăm valorile în Excel
            for col_num, value in enumerate(values_cluburi, start=1):
                cell = ws_cluburi.cell(row=row_num_cluburi, column=col_num)
                cell.value = value
                
                # Font diferit pentru rândurile cu date (nu bold)
                cell.font = openpyxl.styles.Font(bold=False, name="Calibri")
            
            row_num_cluburi += 1

        # Dacă nu avem clasamente de cluburi, adăugăm un mesaj
        if not clasament_cluburi:
            ws_cluburi.cell(row=2, column=1).value = "Nu există clasamente de cluburi disponibile pentru această competiție"
            ws_cluburi.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True, name="Calibri")
        
        # ============= SHEET 3: STATISTICI =============
        ws_statistici = wb.create_sheet(title="Statistici")
        
        # Headers pentru statistici similare cu cele din imagine
        headers_statistici = ["CATEGORIA", "", "", "", "LOC 1", "LOC 2", "LOC 3"]
        
        # Adăugăm headers
        ws_statistici.append(headers_statistici)
        
        # Mergem celulele pentru headerul CATEGORIA
        ws_statistici.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Definim maparea pentru etichetele categoriilor de vârstă
        etichete_varsta = {
            (6, 7): "COPII 1",
            (8, 9): "COPII 2",
            (10, 11): "CADEȚI 1",
            (12, 14): "CADEȚI 2",
            (15, 17): "JUNIORI",
            (18, 20): "TINERET",
            (21, 34): "SENIORI",
            (35, 100): "VETERANI"  # Folosim 100 ca limită superioară pentru veterani
        }
        
        # Funcție pentru a determina eticheta corectă în funcție de vârsta
        def get_eticheta_varsta(varsta_min, varsta_max):
            for (min_range, max_range), eticheta in etichete_varsta.items():
                if varsta_min >= min_range and varsta_max <= max_range:
                    return eticheta
            return "NESPECIFICAT"  # în caz că nu găsim o potrivire
        
        # Grupăm statisticile pe categorii și genuri
        statistici = {}
        
        for clasament in clasamente:
            categorie = clasament.categorie
            sportiv = clasament.sportiv
            
            # Determinăm genul
            gen_display = "FEMININ" if categorie.sex == "F" else "MASCULIN"
            
            # Determinăm categoria de vârstă și eticheta corespunzătoare
            varsta_range = f"{categorie.varsta_min} - {categorie.varsta_max} ANI"
            eticheta = get_eticheta_varsta(categorie.varsta_min, categorie.varsta_max)
            
            # Determinăm categoria de greutate
            cat_kg_display = f"{categorie.categorie_greutate} KG" if categorie.categorie_greutate else "FĂRĂ CATEGORIE"
            
            # Creăm cheia pentru statistici (similar cu formatul din imagine)
            categorie_key = (f"{categorie.proba.nume.upper()} {gen_display}", varsta_range, cat_kg_display, eticheta)
            
            if categorie_key not in statistici:
                statistici[categorie_key] = {1: None, 2: None, 3: None}
            
            # Pentru fiecare locuri 1, 2, 3 salvăm datele sportivului
            sorted_clasamente = sorted(clasamente, key=lambda x: x.puncte)
            
            # Trebuie să găsim locul sportivului curent în categoria sa
            clasamente_aceeasi_categorie = [c for c in clasamente if 
                                           c.categorie.proba.nume == categorie.proba.nume and
                                           c.categorie.sex == categorie.sex and
                                           c.categorie.categorie_greutate == categorie.categorie_greutate and
                                           c.categorie.varsta_min == categorie.varsta_min and
                                           c.categorie.varsta_max == categorie.varsta_max]
            
            clasamente_aceeasi_categorie.sort(key=lambda x: x.puncte)
            
            # Găsim locul sportivului curent
            try:
                loc = clasamente_aceeasi_categorie.index(clasament) + 1
                
                # Salvăm datele doar pentru locurile 1, 2, 3
                if loc <= 3 and statistici[categorie_key][loc] is None:
                    cnp = sportiv.cnp if hasattr(sportiv, 'cnp') else ""
                    club = sportiv.club.nume if sportiv.club else ""
                    
                    statistici[categorie_key][loc] = {
                        'nume': f"{sportiv.nume} {sportiv.prenume}",
                        'cnp': cnp,
                        'club': club
                    }
            except ValueError:
                # Sportivul nu a fost găsit în lista sortată
                pass
        
        # Aplicăm stiluri pentru tabelul de statistici
        for col_idx, header in enumerate(["CATEGORIA", "LOC 1", "LOC 2", "LOC 3"], start=1):
            col_letter = get_column_letter(1 if col_idx == 1 else col_idx + 3)  # Ajustăm pentru poziția reală a coloanelor
            ws_statistici.column_dimensions[col_letter].width = 35  # Lățimea coloanelor
            
            # Styling pentru header
            cell = ws_statistici[f"{col_letter}1"]
            cell.font = openpyxl.styles.Font(bold=True, name="Calibri", italic=True, color="FF0000")  # Roșu
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
        
        # Adăugăm datele pentru statistici
        row_num_statistici = 2
        
        # Sortăm cheile pentru a păstra ordinea categoriilor
        sorted_keys = sorted(statistici.keys())
        
        # Modificarea codului pentru eliminarea bordurilor între rândurile 2-3 și 3-4 pentru fiecare categorie

# ... (cod anterior nemodificat)

        # Înlocuim partea cu bordurile din cod:
        for categorie_key in sorted_keys:
            proba_gen, varsta, kg, eticheta = categorie_key
            
            # Prima coloană - Proba
            proba_cell = ws_statistici.cell(row=row_num_statistici, column=1)
            proba_cell.value = proba_gen
            proba_cell.font = openpyxl.styles.Font(bold=True, name="Calibri", color="FF0000")  # Roșu
            proba_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # A doua coloană - Categoria de vârstă
            varsta_cell = ws_statistici.cell(row=row_num_statistici, column=2)
            varsta_cell.value = varsta
            varsta_cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
            varsta_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # A treia coloană - Categoria de greutate
            kg_cell = ws_statistici.cell(row=row_num_statistici, column=3)
            kg_cell.value = kg
            kg_cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
            kg_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # A patra coloană - Eticheta categoriei de vârstă
            eticheta_cell = ws_statistici.cell(row=row_num_statistici, column=4)
            eticheta_cell.value = eticheta
            eticheta_cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
            eticheta_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Adăugăm datele pentru locurile 1, 2, 3
            for loc in range(1, 4):
                sportiv_data = statistici[categorie_key][loc]
                col_idx = loc + 4  # LOC 1 începe de la coloana 5 (E)
                
                if sportiv_data:
                    # Rândul 1: Numele sportivului
                    nume_cell = ws_statistici.cell(row=row_num_statistici, column=col_idx)
                    nume_cell.value = sportiv_data['nume']
                    nume_cell.font = openpyxl.styles.Font(bold=True, name="Calibri")
                    nume_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    # Rândul 2: CNP-ul sportivului
                    cnp_cell = ws_statistici.cell(row=row_num_statistici + 1, column=col_idx)
                    cnp_cell.value = f"CNP {sportiv_data['cnp']}"
                    cnp_cell.font = openpyxl.styles.Font(bold=True, name="Calibri", color="FF0000")  # Roșu
                    cnp_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    # Rândul 3: Clubul sportivului
                    club_cell = ws_statistici.cell(row=row_num_statistici + 2, column=col_idx)
                    club_cell.value = sportiv_data['club']
                    club_cell.font = openpyxl.styles.Font(bold=True, name="Calibri", color="FF0000")  # Roșu
                    club_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Adăugăm celule goale pentru coloanele de categorie pe rândurile 2 și 3
            for col in range(1, 5):
                for row_offset in range(1, 3):
                    cell = ws_statistici.cell(row=row_num_statistici + row_offset, column=col)
                    cell.value = ""  # Celulă goală
            
            # Aplicăm bordurile doar unde este necesar
            for row in range(row_num_statistici, row_num_statistici + 3):
                for col in range(1, 8):  # 7 coloane în total
                    cell = ws_statistici.cell(row=row, column=col)
                    
                    # Definim bordurile laterale (stânga și dreapta) pentru toate celulele
                    thin_border_sides = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='none'),
                        bottom=openpyxl.styles.Side(style='none')
                    )
                    
                    # Aplicăm borduri complete doar pentru primul și ultimul rând din fiecare grupare de 3 rânduri
                    if row == row_num_statistici:  # Prima linie din grupul de 3
                        border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='none')
                        )
                        cell.border = border
                    elif row == row_num_statistici + 2:  # Ultima linie din grupul de 3
                        border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='none'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.border = border
                    else:  # Linia din mijloc (row_num_statistici + 1)
                        cell.border = thin_border_sides
            
            # După ce am adăugat toate datele pentru o categorie, trecem la rândul următor
            row_num_statistici += 3  # Trecem la următoarea categorie (3 rânduri pentru fiecare categorie)          
        
        # Dacă nu avem statistici, adăugăm un mesaj
        if not statistici:
            ws_statistici.cell(row=2, column=1).value = "Nu există statistici disponibile pentru această competiție"
            ws_statistici.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True, name="Calibri")

        # Pregătim răspunsul HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Clasament Complet {competitie.nume}.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        
        # Salvăm workbook-ul în răspuns
        wb.save(response)
        return response